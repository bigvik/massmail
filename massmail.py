import os
import smtplib
from string import Template
from openpyxl import load_workbook

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr

from secret import config


def send_email(addr_to, msg_subj, msg_text, msg_html):

    msg = MIMEMultipart('alternative')
    msg['From'] = formataddr((str(Header(config['NAME'])), config['FROM']))
    msg['To'] = addr_to
    msg['Subject'] = msg_subj

    msg.attach(MIMEText(msg_text, 'plain'))
    msg.attach(MIMEText(msg_html, 'html'))

    server = smtplib.SMTP_SSL(config['SERVER'], config['PORT'])
    #server.starttls()
    #server.set_debuglevel(True)
    server.login(config['FROM'], config['PASS'])
    server.send_message(msg)
    server.quit()
    print(f'Письмо {addr_to} отправлено')
    
def main():
    wb = load_workbook(filename = 'mailinglist.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_col=3, max_row=3, values_only=True):
        mail_text = f"{row[0]} {row[1]}!\nТестовая рассылка по списку из эксель файла"
        with open('template.html', encoding="utf-8") as tmpl:
            mhtml = Template(tmpl.read())
            mail_html = mhtml.substitute(rich=row[0], name=row[1])

        send_email(row[2], 'Проверка связи', mail_text, mail_html)
    
    
if __name__ == "__main__":
    main()